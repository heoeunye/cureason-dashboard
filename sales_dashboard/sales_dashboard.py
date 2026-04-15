import streamlit as st
import json, os, io, calendar
from datetime import date, timedelta, datetime

DATA_FILE = "sales_data.json"

PLATFORMS = [
    "네이버","카페24","지그재그","에이블리","오늘의집",
    "토스쇼핑","11번가","G마켓","옥션","롯데온",
    "CJ온스타일","SSG","화해","쿠팡","쿠팡(큐리즌 외)",
    "쇼피","큐텐","톡스토어",
]
PLATFORM_MAP = {
    "스마트스토어":"네이버",
    "쿠팡(아원+퓨어픽)":"쿠팡(큐리즌 외)",
    "Shopee":"쇼피",
}

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def month_key(y, m):       return f"{y}-{m:02d}"
def prev_month(y, m):      return (y-1, 12) if m == 1 else (y, m-1)
def platform_total(md, p): return sum(md.get(p, {}).values())
def month_total(md):       return sum(platform_total(md, p) for p in PLATFORMS)
def mom_pct(c, p):         return None if p == 0 else (c-p)/p*100
def fmt_krw(n):            return "—" if n == 0 else f"₩{n:,}"
def fmt_mom(v):
    if v is None: return "—"
    return f"{'+'if v>0 else ''}{v:.1f}%"

def to_date(raw):
    """datetime / date / Excel 시리얼 → date 객체"""
    if isinstance(raw, datetime): return raw.date()
    if isinstance(raw, date):     return raw
    if isinstance(raw, (int, float)):
        return date(1899, 12, 30) + timedelta(days=int(raw))
    return None

def parse_sales_excel(file_bytes):
    import openpyxl
    wb   = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    log  = [f"시트 목록: {wb.sheetnames}"]
    sales, orders = {}, []

    target = [s for s in wb.sheetnames if "매출" in s and "전체" not in s]
    if not target:
        target = [wb.active.title]
    log.append(f"처리할 시트: {target}")

    for sname in target:
        ws   = wb[sname]
        rows = list(ws.iter_rows(values_only=True))

        # 플랫폼 헤더 행 찾기
        ph_row = None
        for i, row in enumerate(rows):
            if row and any(c and "플랫폼" in str(c) and "매출" in str(c) for c in row):
                ph_row = i + 1
                break
        if ph_row is None:
            for i, row in enumerate(rows):
                if sum(1 for c in row if c and str(c).strip() == "날짜") >= 2:
                    ph_row = i
                    break
        if ph_row is None:
            log.append(f"  [{sname}] ❌ 헤더 못 찾음")
            continue

        hdr       = rows[ph_row]
        date_cols = [j for j, c in enumerate(hdr) if c and str(c).strip() == "날짜"]
        date_col  = date_cols[-1] if date_cols else None
        if date_col is None:
            log.append(f"  [{sname}] ❌ 날짜 컬럼 없음")
            continue

        pcols = []
        for j in range(date_col + 1, len(hdr)):
            h = hdr[j]
            if h and str(h).strip() not in ("비고", "None", ""):
                mapped = PLATFORM_MAP.get(str(h).strip(), str(h).strip())
                if mapped in PLATFORMS:
                    pcols.append((j, mapped))

        # 시트 이름에서 기대 월 추출 (예: 매출_2월 → 2)
        MONTH_KR = {"1월":1,"2월":2,"3월":3,"4월":4,"5월":5,"6월":6,
                    "7월":7,"8월":8,"9월":9,"10월":10,"11월":11,"12월":12}
        expected_month = next((v for k,v in MONTH_KR.items() if k in sname), None)

        # 데이터 읽기
        cnt = 0
        for row in rows[ph_row + 1:]:
            if not row or not row[date_col]:
                continue
            d = to_date(row[date_col])
            if not d:
                continue
            # 날짜 연도 보정: 시트명의 월과 날짜의 월이 맞지만 연도가 다른 경우
            if expected_month and d.month == expected_month and d.year != date.today().year:
                try:
                    d = d.replace(year=date.today().year)
                except:
                    pass
            mk = month_key(d.year, d.month)
            sales.setdefault(mk, {})
            for ci, platform in pcols:
                if ci < len(row) and row[ci] and isinstance(row[ci], (int, float)):
                    val = int(row[ci])
                    if val > 0:
                        sales[mk].setdefault(platform, {})[str(d.day)] = val
            cnt += 1

        # 주문 상세 읽기
        for i, row in enumerate(rows):
            strs = [str(c).strip() if c else "" for c in row]
            if "주문형태" in strs and "주문일" in strs:
                for row2 in rows[i + 1:]:
                    if not row2 or not row2[0]:
                        continue
                    try:
                        otype = str(row2[0]).strip()
                        if not otype or otype == "None": continue
                        d2 = to_date(row2[1])
                        if not d2: continue
                        plat  = PLATFORM_MAP.get(str(row2[2]).strip() if row2[2] else "", str(row2[2]).strip() if row2[2] else "")
                        cust  = str(row2[3]).strip() if row2[3] else ""
                        prod  = str(row2[5]).strip() if row2[5] else ""
                        qtype = str(row2[6]).strip() if row2[6] else ""
                        amt   = int(row2[7]) if row2[7] and isinstance(row2[7], (int, float)) else 0
                        orders.append({"date": d2.strftime("%Y-%m-%d"), "platform": plat,
                                       "customer": cust, "product": prod,
                                       "type": qtype, "amount": amt})
                    except:
                        continue
                break

        total = sum(sum(d.values()) for d in sales.get(mk, {}).values()) if sales else 0
        log.append(f"  [{sname}] ✅ {cnt}행 | 월: {[k for k in sales if sname[3:5] in k or True][-1] if sales else '?'}")

    # 중복 주문 제거
    seen, uniq = set(), []
    for o in orders:
        k = (o["date"], o["platform"], o["customer"], o["amount"])
        if k not in seen:
            seen.add(k)
            uniq.append(o)

    log.append(f"최종 감지 월: {sorted(sales.keys())} | 주문 {len(uniq)}건")
    return sales, uniq, log


# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="CUREASON 매출 대시보드", layout="wide")
st.markdown("<style>.block-container{padding-top:1.4rem}</style>", unsafe_allow_html=True)

data = load_data()

st.markdown("### CUREASON 매출 대시보드")
today = date.today()
c1, c2, _ = st.columns([1, 1, 3])
sel_year  = c1.number_input("연도", 2020, 2030, today.year, step=1)
sel_month = c2.selectbox("월", range(1, 13), index=today.month-1,
                          format_func=lambda x: f"{x}월")
st.divider()

cur_key  = month_key(sel_year, sel_month)
py, pm_  = prev_month(sel_year, sel_month)
prev_key = month_key(py, pm_)
cur_data  = data.get(cur_key, {})
prev_data = data.get(prev_key, {})

tab_sum, tab_input, tab_import, tab_orders = st.tabs(
    ["📊 월별 요약", "✏️ 일자별 입력", "📥 엑셀 가져오기", "📋 주문 상세"]
)

# ══ 탭1: 월별 요약
with tab_sum:
    ct  = month_total(cur_data)
    pt  = month_total(prev_data)
    pct = mom_pct(ct, pt)
    c1, c2, c3 = st.columns(3)
    c1.metric("당월 합계", fmt_krw(ct))
    c2.metric("전월 합계", fmt_krw(pt))
    c3.metric("전월비", fmt_mom(pct), delta=f"{pct:.1f}%" if pct else None)
    st.markdown(f"---\n**플랫폼별 상세** — {sel_year}.{sel_month:02d} vs {py}.{pm_:02d}")
    for col, label in zip(st.columns([3,2,2,2]), ["플랫폼","당월","전월","전월비"]):
        col.markdown(f"**{label}**")
    for p in PLATFORMS:
        cp  = platform_total(cur_data, p)
        pp  = platform_total(prev_data, p)
        if cp == 0 and pp == 0: continue
        pp_ = mom_pct(cp, pp)
        r   = st.columns([3,2,2,2])
        r[0].write(p)
        r[1].write(fmt_krw(cp))
        r[2].write(fmt_krw(pp) if pp else "—")
        if pp_ is None:   r[3].write("—")
        elif pp_ > 0:     r[3].markdown(f":green[{fmt_mom(pp_)}]")
        else:             r[3].markdown(f":red[{fmt_mom(pp_)}]")

# ══ 탭2: 일자별 입력
with tab_input:
    dim     = calendar.monthrange(sel_year, sel_month)[1]
    sel_day = st.selectbox("날짜", range(1, dim+1),
                index=min(today.day, dim)-1,
                format_func=lambda d: f"{sel_year}.{sel_month:02d}.{d:02d}")
    st.markdown(f"**{sel_year}.{sel_month:02d}.{sel_day:02d} 매출 입력**")
    input_vals = {}
    for chunk in [PLATFORMS[i:i+3] for i in range(0, len(PLATFORMS), 3)]:
        cols = st.columns(3)
        for i, p in enumerate(chunk):
            existing = int(cur_data.get(p, {}).get(str(sel_day), 0))
            input_vals[p] = cols[i].number_input(p, min_value=0, value=existing, step=100, key=f"inp_{p}")
    st.markdown(f"**일 합계: {fmt_krw(sum(input_vals.values()))}**")
    if st.button("💾 저장하기", type="primary", use_container_width=True):
        data.setdefault(cur_key, {})
        for p, v in input_vals.items():
            data[cur_key].setdefault(p, {})
            if v > 0: data[cur_key][p][str(sel_day)] = v
            else:     data[cur_key][p].pop(str(sel_day), None)
        save_data(data)
        st.success("저장 완료!")
        st.rerun()

# ══ 탭3: 엑셀 가져오기
with tab_import:
    st.markdown("#### 매출표 엑셀 파일 가져오기")
    st.caption("파일 안의 모든 월별 시트를 한 번에 읽어옵니다.")
    uploaded = st.file_uploader("xlsx 파일 선택", type=["xlsx"])

    if uploaded:
        with st.spinner("전체 시트 파싱 중..."):
            try:
                parsed_sales, parsed_orders, log = parse_sales_excel(uploaded.read())
            except Exception as e:
                import traceback
                st.error(f"오류: {e}")
                st.code(traceback.format_exc())
                parsed_sales, parsed_orders, log = {}, [], []

        with st.expander("🔍 파싱 로그", expanded=not bool(parsed_sales)):
            for line in log:
                st.text(line)

        if not parsed_sales:
            st.error("파싱 실패 — 로그를 확인해주세요.")
        else:
            st.success(f"파싱 완료! {sorted(parsed_sales.keys())}")
            for mk in sorted(parsed_sales.keys()):
                md    = parsed_sales[mk]
                total = month_total(md)
                st.markdown(f"**{mk} — {fmt_krw(total)}**")
                active = [(p, platform_total(md, p)) for p in PLATFORMS if platform_total(md, p) > 0]
                cols = st.columns(4)
                for i, (p, amt) in enumerate(active):
                    cols[i%4].write(f"{p}: {fmt_krw(amt)}")
                st.markdown("---")

            if parsed_orders:
                st.markdown(f"주문 상세 {len(parsed_orders)}건")

            overwrite = st.checkbox("⚠️ 기존 데이터 덮어쓰기", value=True)
            if st.button("📥 가져오기 확정", type="primary", use_container_width=True):
                for mk, md in parsed_sales.items():
                    if mk not in data or overwrite:
                        data[mk] = md
                    else:
                        for platform, days in md.items():
                            data[mk].setdefault(platform, {}).update(days)
                if parsed_orders:
                    if overwrite:
                        data["orders"] = parsed_orders
                    else:
                        existing = {(o["date"],o["platform"],o["customer"]) for o in data.get("orders",[])}
                        data.setdefault("orders",[]).extend(
                            o for o in parsed_orders
                            if (o["date"],o["platform"],o["customer"]) not in existing)
                save_data(data)
                st.success("✅ 저장 완료!")
                st.rerun()

# ══ 탭4: 주문 상세
with tab_orders:
    st.markdown("#### 주문 상세")
    all_orders = data.get("orders", [])
    if not all_orders:
        st.info("주문 데이터 없음 — 엑셀 가져오기로 불러오세요.")
    else:
        fc1, fc2 = st.columns(2)
        f_month    = fc1.text_input("월 필터 (예: 2026-04)")
        f_platform = fc2.selectbox("플랫폼", ["전체"] + PLATFORMS)
        filtered = all_orders
        if f_month:    filtered = [o for o in filtered if o["date"].startswith(f_month)]
        if f_platform != "전체": filtered = [o for o in filtered if o["platform"] == f_platform]
        st.caption(f"총 {len(filtered)}건 | 합계 {fmt_krw(sum(o['amount'] for o in filtered))}")
        import pandas as pd
        df = pd.DataFrame(filtered)[["date","platform","customer","product","type","amount"]]
        df.columns = ["날짜","플랫폼","주문자","상품","단/복수","공급가"]
        st.dataframe(df, use_container_width=True, hide_index=True)